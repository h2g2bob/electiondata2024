import openpyxl
import sqlite3

FILENAME = "Constituencies by county and LA revised proposals England.xlsx"

def main():
    con = sqlite3.connect("../data.sqlite3")
    with con:
        con.execute("drop table if exists bce_proposed")
        con.execute(
            """
            create table bce_proposed(
                -- ward id is NOT unique!!
                -- (eg: if a ward is split between multiple constituencies)
                ward_id text not null,
                ward_name text not null,
                local_name text not null,
                electorate integer,
                initial text not null,
                revised text not null)
            """
        )
        import_data(con)

def import_data(con):
    wb = openpyxl.load_workbook(filename=FILENAME, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = iter([cell.value for cell in row] for row in sheet.iter_rows())

        header = next(rows)
        assert header[0].startswith("Local authority electorates")
        consume_to_empty_row(rows)
        header = next(rows)
        assert header[0].startswith("Revised proposed constituency electorates by local authority")
        consume_to_empty_row(rows)
        header = next(rows)
        assert header[0].startswith("Revised proposed constituency electorates by initial proposed constituency")
        consume_to_empty_row(rows)
        while True:
            try:
                header = next(rows)
            except StopIteration:
                break
            assert header[0].startswith("Electorate by ward for")
            read_local_authority_data(con, rows)

def consume_to_empty_row(row_iter):
    for row in row_iter:
        if row[0] is None:
            assert all(cell is None for cell in row)
            return
    raise StopIteration()

def read_local_authority_data(con, row_iter):
    header = next(row_iter)
    assert header == ["Local authority", "Ward", "ONS code", "Electorate", "Initial proposed constituency", "Revised proposed constituency"]
    for row in row_iter:
        if row[0] is None:
            assert all(cell is None for cell in row)
            return
        con.execute("insert into bce_proposed(local_name, ward_name, ward_id, electorate, initial, revised) values (?, ?, ?, ?, ?, ?)", row)
    return

if __name__ == "__main__":
    main()
